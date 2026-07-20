from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def cell_value(cell):
    value = cell.value
    if value is None:
        return None
    return {
        "coordinate": cell.coordinate,
        "value": value,
        "type": cell.data_type,
    }


def summarize(path: Path, selected_sheet: str | None = None) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    try:
        for worksheet in workbook.worksheets:
            if selected_sheet and worksheet.title != selected_sheet:
                continue
            populated = []
            formulas = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    item = cell_value(cell)
                    if item is None:
                        continue
                    if cell.data_type == "f":
                        formulas += 1
                    if len(populated) < 120:
                        populated.append(item)
            sheets.append(
                {
                    "name": worksheet.title,
                    "rows": worksheet.max_row,
                    "columns": worksheet.max_column,
                    "formula_count": formulas,
                    "sample": populated,
                }
            )
    finally:
        workbook.close()
    return {"workbook": path.name, "sheets": sheets}


if __name__ == "__main__":
    selected = None
    arguments = sys.argv[1:]
    if "--sheet" in arguments:
        index = arguments.index("--sheet")
        selected = arguments[index + 1]
        del arguments[index : index + 2]
    for argument in arguments:
        print(json.dumps(summarize(Path(argument), selected), ensure_ascii=False, indent=2))
